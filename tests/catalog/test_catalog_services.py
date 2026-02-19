from io import BytesIO

import openpyxl
import pytest
from django.utils import timezone

from catalog.models import Brand, Category, Product
from catalog.services import (
    IMPORT_COLUMNS,
    _unique_slug,
    export_products_to_excel,
    import_products_from_excel,
)


def _build_excel_file(rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(IMPORT_COLUMNS)
    for row in rows:
        ws.append(row)
    content = BytesIO()
    wb.save(content)
    content.seek(0)
    return content


@pytest.mark.django_db
class TestCatalogServices:
    def test_import_creates_product_category_and_brand(self, enterprise):
        file = _build_excel_file(
            [
                (
                    "Switch 8 ports",
                    "SW-8P",
                    "123456",
                    "Reseau",
                    "Cisco",
                    "10000",
                    "15000",
                    "Produit test",
                    "oui",
                )
            ]
        )

        result = import_products_from_excel(file, enterprise=enterprise)

        assert result == {
            "created": 1,
            "updated": 0,
            "errors": 0,
            "error_details": [],
        }
        product = Product.objects.get(sku="SW-8P")
        assert product.name == "Switch 8 ports"
        assert product.category.name == "Reseau"
        assert product.brand.name == "Cisco"
        assert product.is_active is True

    def test_import_updates_existing_product_and_generates_unique_slug(self, enterprise):
        category = Category.objects.create(
            enterprise=enterprise,
            name="Reseau",
            slug="reseau",
        )
        Product.objects.create(
            enterprise=enterprise,
            category=category,
            brand=None,
            name="Routeur Pro",
            slug="routeur-pro",
            sku="ROUT-01",
            selling_price="30000.00",
            cost_price="20000.00",
        )
        target = Product.objects.create(
            enterprise=enterprise,
            category=category,
            brand=None,
            name="Ancien nom",
            slug="ancien-nom",
            sku="ROUT-02",
            selling_price="12000.00",
            cost_price="10000.00",
        )

        file = _build_excel_file(
            [
                (
                    "Routeur Pro",
                    "ROUT-02",
                    "",
                    "Reseau",
                    "",
                    "21000",
                    "33000",
                    "",
                    1,
                )
            ]
        )

        result = import_products_from_excel(file, enterprise=enterprise)
        target.refresh_from_db()

        assert result["created"] == 0
        assert result["updated"] == 1
        assert target.name == "Routeur Pro"
        assert target.slug == "routeur-pro-1"

    def test_import_returns_error_when_category_is_missing(self, enterprise):
        file = _build_excel_file(
            [
                (
                    "Switch 16 ports",
                    "SW-16P",
                    "987654",
                    "",
                    "Cisco",
                    "10000",
                    "16000",
                    "",
                    "non",
                )
            ]
        )

        result = import_products_from_excel(file, enterprise=enterprise)

        assert result["created"] == 0
        assert result["updated"] == 0
        assert result["errors"] == 1
        assert len(result["error_details"]) == 1
        assert "Ligne 2" in result["error_details"][0]
        assert "categorie est obligatoire" in result["error_details"][0].lower()

    def test_export_products_to_excel_generates_expected_rows(self, enterprise):
        category = Category.objects.create(
            enterprise=enterprise,
            name="Reseau",
            slug="reseau",
        )
        brand = Brand.objects.create(
            enterprise=enterprise,
            name="Cisco",
            slug="cisco",
        )
        product = Product.objects.create(
            enterprise=enterprise,
            category=category,
            brand=brand,
            name="Switch 24 ports",
            slug="switch-24-ports",
            sku="SW-24P",
            barcode="ABC123",
            description="Modele entreprise",
            cost_price="25000.00",
            selling_price="40000.00",
            is_active=True,
        )
        product.created_at = timezone.now()
        product.save(update_fields=["created_at"])

        response = export_products_to_excel(Product.objects.filter(pk=product.pk))
        wb = openpyxl.load_workbook(BytesIO(response.content), data_only=True)
        ws = wb.active

        assert (
            response["Content-Type"]
            == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        assert response["Content-Disposition"] == 'attachment; filename="produits_export.xlsx"'
        assert ws.cell(row=1, column=1).value == "Nom"
        assert ws.cell(row=1, column=12).value == "Date de creation"
        assert ws.cell(row=2, column=1).value == "Switch 24 ports"
        assert ws.cell(row=2, column=2).value == "SW-24P"
        assert ws.cell(row=2, column=4).value == "Reseau"
        assert ws.cell(row=2, column=5).value == "Cisco"
        assert ws.cell(row=2, column=11).value == "Oui"
        assert ws.cell(row=2, column=12).value == product.created_at.strftime("%Y-%m-%d %H:%M")
        wb.close()

    def test_unique_slug_appends_suffix_on_conflict(self, enterprise):
        Category.objects.create(
            enterprise=enterprise,
            name="Reseau",
            slug="reseau",
        )

        slug = _unique_slug(Category, "Reseau", enterprise=enterprise)

        assert slug == "reseau-1"

    def test_unique_slug_uses_fallback_when_slugify_returns_empty(self, enterprise):
        slug = _unique_slug(Category, "!!!", enterprise=enterprise)
        Category.objects.create(enterprise=enterprise, name="Divers", slug=slug)

        next_slug = _unique_slug(Category, "***", enterprise=enterprise)

        assert slug == "item"
        assert next_slug == "item-1"
