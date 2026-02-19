"""
Service functions for the catalog app.

Handles bulk import from Excel and export to Excel using openpyxl.
"""
import logging
from decimal import Decimal, InvalidOperation
from io import BytesIO

from django.http import HttpResponse
from django.utils.text import slugify

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .models import Brand, Category, Product

logger = logging.getLogger("boutique")

# ---------------------------------------------------------------------------
# Expected column order for the import spreadsheet
# ---------------------------------------------------------------------------
IMPORT_COLUMNS = [
    "nom",           # A - name
    "sku",           # B - sku (unique reference)
    "code_barres",   # C - barcode
    "categorie",     # D - category name
    "marque",        # E - brand name
    "prix_achat",    # F - cost_price
    "prix_vente",    # G - selling_price
    "description",   # H - description
    "actif",         # I - is_active (oui/non or 1/0)
]


# =========================================================================
# IMPORT
# =========================================================================

def import_products_from_excel(file, enterprise=None) -> dict:
    """
    Import products from an uploaded Excel (.xlsx) file.

    Expected columns (first row is header):
        nom | sku | code_barres | categorie | marque | prix_achat | prix_vente | description | actif

    Returns a dict with counts::

        {"created": int, "updated": int, "errors": int, "error_details": list[str]}
    """
    wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
    ws = wb.active

    created = 0
    updated = 0
    errors = 0
    error_details: list[str] = []

    rows = ws.iter_rows(min_row=2, values_only=True)  # skip header
    for row_idx, row in enumerate(rows, start=2):
        try:
            # Unpack row (pad with None when columns are missing)
            padded = list(row) + [None] * (len(IMPORT_COLUMNS) - len(row))
            (
                name,
                sku,
                barcode,
                category_name,
                brand_name,
                cost_price,
                selling_price,
                description,
                is_active_raw,
            ) = padded[:9]

            # ----- Validation -----
            if not name or not sku:
                raise ValueError("Le nom et le SKU sont obligatoires.")

            if not selling_price:
                raise ValueError("Le prix de vente est obligatoire.")

            name = str(name).strip()
            sku = str(sku).strip()
            barcode = str(barcode).strip() if barcode else ""

            # Category (create if it does not exist)
            category = None
            if category_name:
                category_name = str(category_name).strip()
                cat_lookup = {"name__iexact": category_name}
                cat_defaults = {
                    "name": category_name,
                    "slug": _unique_slug(Category, category_name, enterprise=enterprise),
                }
                if enterprise:
                    cat_lookup["enterprise"] = enterprise
                    cat_defaults["enterprise"] = enterprise
                category, _ = Category.objects.get_or_create(
                    **cat_lookup,
                    defaults=cat_defaults,
                )

            if category is None:
                raise ValueError("La categorie est obligatoire.")

            # Brand (create if it does not exist, optional)
            brand = None
            if brand_name:
                brand_name = str(brand_name).strip()
                brand_lookup = {"name__iexact": brand_name}
                brand_defaults = {
                    "name": brand_name,
                    "slug": _unique_slug(Brand, brand_name, enterprise=enterprise),
                }
                if enterprise:
                    brand_lookup["enterprise"] = enterprise
                    brand_defaults["enterprise"] = enterprise
                brand, _ = Brand.objects.get_or_create(
                    **brand_lookup,
                    defaults=brand_defaults,
                )

            # Prices
            try:
                cost_price = Decimal(str(cost_price)) if cost_price else Decimal("0.00")
                selling_price = Decimal(str(selling_price))
            except (InvalidOperation, TypeError):
                raise ValueError("Prix invalides.")

            # is_active
            if is_active_raw is None:
                is_active = True
            elif isinstance(is_active_raw, bool):
                is_active = is_active_raw
            else:
                is_active = str(is_active_raw).strip().lower() in (
                    "1", "oui", "true", "yes", "vrai",
                )

            description = str(description).strip() if description else ""

            # ----- Create or update -----
            product_lookup = {"sku": sku}
            product_defaults = {
                "name": name,
                "slug": _unique_slug(Product, name, enterprise=enterprise, exclude_sku=sku),
                "barcode": barcode,
                "category": category,
                "brand": brand,
                "cost_price": cost_price,
                "selling_price": selling_price,
                "description": description,
                "is_active": is_active,
            }
            if enterprise:
                product_lookup["enterprise"] = enterprise
                product_defaults["enterprise"] = enterprise
            product, was_created = Product.objects.update_or_create(
                **product_lookup,
                defaults=product_defaults,
            )

            if was_created:
                created += 1
            else:
                updated += 1

        except Exception as exc:
            errors += 1
            detail = f"Ligne {row_idx}: {exc}"
            error_details.append(detail)
            logger.warning("Import produit - %s", detail)

    wb.close()

    logger.info(
        "Import produits termine: %d cree(s), %d mis a jour, %d erreur(s).",
        created, updated, errors,
    )

    return {
        "created": created,
        "updated": updated,
        "errors": errors,
        "error_details": error_details,
    }


# =========================================================================
# EXPORT
# =========================================================================

def export_products_to_excel(queryset) -> HttpResponse:
    """
    Export a queryset of products to an Excel (.xlsx) file returned as an
    ``HttpResponse`` suitable for direct download.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Produits"

    # ----- Header row -----
    headers = [
        "Nom",
        "SKU",
        "Code-barres",
        "Categorie",
        "Marque",
        "Prix d'achat",
        "Prix de vente",
        "Marge",
        "Marge %",
        "Description",
        "Actif",
        "Date de creation",
    ]

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # ----- Data rows -----
    for row_num, product in enumerate(queryset.iterator(), start=2):
        ws.cell(row=row_num, column=1, value=product.name)
        ws.cell(row=row_num, column=2, value=product.sku)
        ws.cell(row=row_num, column=3, value=product.barcode)
        ws.cell(row=row_num, column=4, value=product.category.name if product.category else "")
        ws.cell(row=row_num, column=5, value=product.brand.name if product.brand else "")
        ws.cell(row=row_num, column=6, value=float(product.cost_price))
        ws.cell(row=row_num, column=7, value=float(product.selling_price))
        ws.cell(row=row_num, column=8, value=float(product.margin))
        ws.cell(row=row_num, column=9, value=float(product.margin_percent))
        ws.cell(row=row_num, column=10, value=product.description)
        ws.cell(row=row_num, column=11, value="Oui" if product.is_active else "Non")
        ws.cell(row=row_num, column=12, value=product.created_at.strftime("%Y-%m-%d %H:%M"))

    # ----- Auto-size columns -----
    for col_num in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_num)
        max_length = len(str(headers[col_num - 1]))
        for row in ws.iter_rows(min_row=2, min_col=col_num, max_col=col_num):
            for cell in row:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_length + 4, 50)

    # ----- Build response -----
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="produits_export.xlsx"'
    return response


# =========================================================================
# Helpers
# =========================================================================

def _unique_slug(model_class, name: str, enterprise=None, exclude_sku: str | None = None) -> str:
    """
    Generate a unique slug for the given model class based on *name*,
    scoped to the given enterprise.

    If a product with the same slug already exists (and is not the one
    identified by *exclude_sku*), append a numeric suffix.
    """
    base_slug = slugify(name) or "item"
    slug = base_slug
    counter = 1

    while True:
        qs = model_class.objects.filter(slug=slug)
        if enterprise:
            qs = qs.filter(enterprise=enterprise)
        if exclude_sku and model_class is Product:
            qs = qs.exclude(sku=exclude_sku)
        if not qs.exists():
            return slug
        slug = f"{base_slug}-{counter}"
        counter += 1
